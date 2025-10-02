"""Data quality analysis for Excel files"""

import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional
import logging
import re

logger = logging.getLogger(__name__)


class DataQualityAnalyser:
    """Analyzes data quality in Excel sheets"""

    def __init__(self):
        self.workbook = None

    def analyse(self, excel_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Analyze data quality across sheets

        Args:
            excel_path: Path to Excel file
            sheet_name: Optional specific sheet to analyze (None = all sheets)

        Returns:
            Dictionary containing data quality analysis results
        """
        logger.info(f"Analyzing data quality of: {excel_path}")

        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)

        results = {
            "file_path": excel_path,
            "sheets_analyzed": [],
            "overall_quality_score": 0.0,
            "issues_summary": {
                "total_issues": 0,
                "critical_issues": 0,
                "warnings": 0
            }
        }

        # Determine which sheets to analyze
        sheets_to_analyze = [sheet_name] if sheet_name else self.workbook.sheetnames

        quality_scores = []

        for sheet in sheets_to_analyze:
            if sheet not in self.workbook.sheetnames:
                logger.warning(f"Sheet '{sheet}' not found, skipping")
                continue

            sheet_analysis = self._analyse_sheet(excel_path, sheet)
            results["sheets_analyzed"].append(sheet_analysis)
            quality_scores.append(sheet_analysis["quality_score"])

            # Aggregate issues
            results["issues_summary"]["total_issues"] += sheet_analysis["issues"]["total"]
            results["issues_summary"]["critical_issues"] += sheet_analysis["issues"]["critical"]
            results["issues_summary"]["warnings"] += sheet_analysis["issues"]["warnings"]

        # Calculate overall quality score
        if quality_scores:
            results["overall_quality_score"] = sum(quality_scores) / len(quality_scores)

        logger.info(f"Data quality analysis complete. Overall score: {results['overall_quality_score']:.2f}")
        return results

    def _analyse_sheet(self, excel_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        Analyze data quality for a single sheet

        Args:
            excel_path: Path to Excel file
            sheet_name: Name of the sheet

        Returns:
            Dictionary with sheet quality analysis
        """
        try:
            # Read sheet as DataFrame
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
        except Exception as e:
            logger.error(f"Error reading sheet '{sheet_name}': {str(e)}")
            return {
                "sheet_name": sheet_name,
                "error": str(e),
                "quality_score": 0.0
            }

        analysis = {
            "sheet_name": sheet_name,
            "row_count": len(df),
            "column_count": len(df.columns),
            "missing_values": {},
            "duplicate_rows": 0,
            "data_types": {},
            "issues": {
                "total": 0,
                "critical": 0,
                "warnings": 0,
                "details": []
            },
            "quality_score": 100.0  # Start with perfect score, deduct for issues
        }

        # Check for missing values
        missing_analysis = self._check_missing_values(df)
        analysis["missing_values"] = missing_analysis

        # Check for duplicates
        duplicates_count = df.duplicated().sum()
        analysis["duplicate_rows"] = int(duplicates_count)

        if duplicates_count > 0:
            analysis["issues"]["warnings"] += 1
            analysis["issues"]["total"] += 1
            analysis["issues"]["details"].append({
                "type": "warning",
                "message": f"Found {duplicates_count} duplicate rows"
            })
            analysis["quality_score"] -= min(10, duplicates_count * 0.5)

        # Analyze data types
        dtype_analysis = self._analyse_data_types(df)
        analysis["data_types"] = dtype_analysis

        # Check for inconsistent data types
        inconsistent_types = self._check_inconsistent_types(df)
        if inconsistent_types:
            analysis["issues"]["warnings"] += len(inconsistent_types)
            analysis["issues"]["total"] += len(inconsistent_types)
            for col, issue in inconsistent_types.items():
                analysis["issues"]["details"].append({
                    "type": "warning",
                    "message": f"Column '{col}': {issue}"
                })
            analysis["quality_score"] -= min(15, len(inconsistent_types) * 3)

        # Check for missing values issues
        critical_missing = [col for col, pct in missing_analysis.items() if pct > 50]
        if critical_missing:
            analysis["issues"]["critical"] += len(critical_missing)
            analysis["issues"]["total"] += len(critical_missing)
            for col in critical_missing:
                analysis["issues"]["details"].append({
                    "type": "critical",
                    "message": f"Column '{col}' has >50% missing values ({missing_analysis[col]:.1f}%)"
                })
            analysis["quality_score"] -= min(30, len(critical_missing) * 10)

        # Check for empty columns
        empty_columns = [col for col in df.columns if df[col].isna().all()]
        if empty_columns:
            analysis["issues"]["warnings"] += len(empty_columns)
            analysis["issues"]["total"] += len(empty_columns)
            analysis["issues"]["details"].append({
                "type": "warning",
                "message": f"Empty columns found: {', '.join(empty_columns)}"
            })
            analysis["quality_score"] -= min(10, len(empty_columns) * 2)

        # Check for unnamed columns
        unnamed_cols = [col for col in df.columns if 'Unnamed' in str(col)]
        if unnamed_cols:
            analysis["issues"]["warnings"] += len(unnamed_cols)
            analysis["issues"]["total"] += len(unnamed_cols)
            analysis["issues"]["details"].append({
                "type": "warning",
                "message": f"Unnamed columns found: {len(unnamed_cols)}"
            })
            analysis["quality_score"] -= min(5, len(unnamed_cols) * 1)

        # Check for whitespace issues
        whitespace_issues = self._check_whitespace_issues(df)
        if whitespace_issues:
            analysis["issues"]["warnings"] += len(whitespace_issues)
            analysis["issues"]["total"] += len(whitespace_issues)
            for col, count in whitespace_issues.items():
                analysis["issues"]["details"].append({
                    "type": "warning",
                    "message": f"Column '{col}': {count} cells with leading/trailing whitespace"
                })
            analysis["quality_score"] -= min(5, len(whitespace_issues) * 1)

        # Ensure score doesn't go below 0
        analysis["quality_score"] = max(0, analysis["quality_score"])

        return analysis

    def _check_missing_values(self, df: pd.DataFrame) -> Dict[str, float]:
        """Check for missing values and calculate percentages"""
        missing_pct = {}
        for col in df.columns:
            missing_count = df[col].isna().sum()
            if missing_count > 0:
                pct = (missing_count / len(df)) * 100
                missing_pct[col] = round(pct, 2)
        return missing_pct

    def _analyse_data_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """Analyze data types of columns"""
        dtype_map = {}
        for col in df.columns:
            dtype_map[col] = str(df[col].dtype)
        return dtype_map

    def _check_inconsistent_types(self, df: pd.DataFrame) -> Dict[str, str]:
        """Check for columns with mixed data types"""
        inconsistent = {}

        for col in df.columns:
            # Skip if all NaN
            if df[col].isna().all():
                continue

            # Get non-null values
            non_null = df[col].dropna()

            if len(non_null) == 0:
                continue

            # Check for mixed types
            types = set(type(v).__name__ for v in non_null)

            # If more than one type (excluding NoneType), flag it
            if len(types) > 1:
                inconsistent[col] = f"Mixed types: {', '.join(types)}"

            # Check for numeric columns with text values
            if df[col].dtype == 'object':
                # Try to convert to numeric
                try:
                    numeric_vals = pd.to_numeric(non_null, errors='coerce')
                    non_numeric_count = numeric_vals.isna().sum()

                    if non_numeric_count > 0 and non_numeric_count < len(non_null):
                        pct = (non_numeric_count / len(non_null)) * 100
                        if pct < 90:  # If less than 90% are non-numeric
                            inconsistent[col] = f"{non_numeric_count} non-numeric values in mostly numeric column"
                except:
                    pass

        return inconsistent

    def _check_whitespace_issues(self, df: pd.DataFrame) -> Dict[str, int]:
        """Check for leading/trailing whitespace in text columns"""
        whitespace_issues = {}

        for col in df.columns:
            if df[col].dtype == 'object':
                # Check for whitespace
                whitespace_count = 0
                for val in df[col].dropna():
                    if isinstance(val, str) and (val != val.strip()):
                        whitespace_count += 1

                if whitespace_count > 0:
                    whitespace_issues[col] = whitespace_count

        return whitespace_issues

    def get_quality_summary(self, excel_path: str, sheet_name: Optional[str] = None) -> str:
        """
        Get a text summary of data quality

        Args:
            excel_path: Path to Excel file
            sheet_name: Optional specific sheet

        Returns:
            Formatted text summary
        """
        analysis = self.analyse(excel_path, sheet_name)

        summary = []
        summary.append("=" * 70)
        summary.append("DATA QUALITY ANALYSIS")
        summary.append("=" * 70)
        summary.append(f"\nFile: {excel_path}")
        summary.append(f"Overall Quality Score: {analysis['overall_quality_score']:.1f}/100")
        summary.append(f"\nTotal Issues: {analysis['issues_summary']['total_issues']}")
        summary.append(f"  Critical: {analysis['issues_summary']['critical_issues']}")
        summary.append(f"  Warnings: {analysis['issues_summary']['warnings']}")

        for sheet_analysis in analysis['sheets_analyzed']:
            summary.append("\n" + "-" * 70)
            summary.append(f"Sheet: {sheet_analysis['sheet_name']}")
            summary.append("-" * 70)

            if 'error' in sheet_analysis:
                summary.append(f"  ERROR: {sheet_analysis['error']}")
                continue

            summary.append(f"  Quality Score: {sheet_analysis['quality_score']:.1f}/100")
            summary.append(f"  Rows: {sheet_analysis['row_count']}, Columns: {sheet_analysis['column_count']}")
            summary.append(f"  Duplicate Rows: {sheet_analysis['duplicate_rows']}")

            if sheet_analysis['missing_values']:
                summary.append(f"\n  Missing Values:")
                for col, pct in sheet_analysis['missing_values'].items():
                    summary.append(f"    - {col}: {pct}%")

            if sheet_analysis['issues']['details']:
                summary.append(f"\n  Issues Found:")
                for issue in sheet_analysis['issues']['details']:
                    summary.append(f"    [{issue['type'].upper()}] {issue['message']}")

        summary.append("\n" + "=" * 70)

        return "\n".join(summary)

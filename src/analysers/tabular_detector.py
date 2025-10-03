"""Tabular structure detection for Excel sheets"""

import pandas as pd
import openpyxl
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
import logging

logger = logging.getLogger(__name__)


class TabularDetector:
    """Detects if Excel sheets contain tabular structures, including multiple tables per sheet"""

    def __init__(self, min_table_rows: int = 3, min_table_cols: int = 2, blank_row_threshold: int = 2, blank_col_threshold: int = 1):
        """
        Initialize TabularDetector

        Args:
            min_table_rows: Minimum rows for a valid table
            min_table_cols: Minimum columns for a valid table
            blank_row_threshold: Number of consecutive blank rows to consider as separator
            blank_col_threshold: Number of consecutive blank columns to consider as separator
        """
        self.workbook = None
        self.min_table_rows = min_table_rows
        self.min_table_cols = min_table_cols
        self.blank_row_threshold = blank_row_threshold
        self.blank_col_threshold = blank_col_threshold

    def analyse(self, excel_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Detect tabular structures in Excel sheets (supports multiple tables per sheet)

        Args:
            excel_path: Path to Excel file
            sheet_name: Optional specific sheet to analyze (None = all sheets)

        Returns:
            Dictionary containing tabular structure detection results
        """
        logger.info(f"Detecting tabular structures in: {excel_path}")

        self.workbook = openpyxl.load_workbook(excel_path, data_only=False)

        results = {
            "file_path": excel_path,
            "sheets_analyzed": [],
            "summary": {
                "total_sheets": 0,
                "total_tables": 0,
                "sheets_with_tables": 0,
                "sheets_without_tables": 0
            }
        }

        # Determine which sheets to analyze
        sheets_to_analyze = [sheet_name] if sheet_name else self.workbook.sheetnames

        for sheet in sheets_to_analyze:
            if sheet not in self.workbook.sheetnames:
                logger.warning(f"Sheet '{sheet}' not found, skipping")
                continue

            sheet_analysis = self._detect_tabular_structure(excel_path, sheet)
            results["sheets_analyzed"].append(sheet_analysis)

            # Update summary
            results["summary"]["total_sheets"] += 1
            results["summary"]["total_tables"] += sheet_analysis.get("tables_found", 0)

            if sheet_analysis.get("tables_found", 0) > 0:
                # Check if any table is actually tabular
                has_tabular = any(t.get("is_tabular", False) for t in sheet_analysis.get("tables", []))
                if has_tabular:
                    results["summary"]["sheets_with_tables"] += 1
                else:
                    results["summary"]["sheets_without_tables"] += 1
            else:
                results["summary"]["sheets_without_tables"] += 1

        logger.info(
            f"Tabular detection complete: {results['summary']['total_tables']} tables found "
            f"across {results['summary']['sheets_with_tables']} sheets"
        )
        return results

    def _detect_tabular_structure(self, excel_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        Detect if a sheet has tabular structures (can be multiple tables)

        Args:
            excel_path: Path to Excel file
            sheet_name: Name of the sheet

        Returns:
            Dictionary with detection results for all tables in the sheet
        """
        try:
            # Read sheet with openpyxl for structure analysis
            sheet = self.workbook[sheet_name]

        except Exception as e:
            logger.error(f"Error analyzing sheet '{sheet_name}': {str(e)}")
            return {
                "sheet_name": sheet_name,
                "error": str(e),
                "tables_found": 0,
                "tables": []
            }

        # Find all data regions in the sheet
        regions = self._find_data_regions(sheet)

        if not regions:
            logger.info(f"No data regions found in sheet '{sheet_name}'")
            return {
                "sheet_name": sheet_name,
                "tables_found": 0,
                "tables": [],
                "message": "No data regions detected"
            }

        # Analyze each region
        all_tables = []
        table_id = 1

        for region in regions:
            logger.info(f"Analyzing region: {region}")

            # Extract data for this region
            try:
                # Read the specific region using openpyxl coordinates
                df = pd.read_excel(
                    excel_path,
                    sheet_name=sheet_name,
                    header=None,  # We'll detect header ourselves
                    skiprows=region['start_row'] - 1,
                    nrows=region['end_row'] - region['start_row'],
                    usecols=range(region['start_col'] - 1, region['end_col'] - 1)
                )

                # Detect tables within this region
                tables_in_region = self._detect_table_boundaries(df, region)

                for table_def in tables_in_region:
                    # Analyze this specific table
                    table_analysis = self._analyze_single_table(df, table_def, sheet, table_id)
                    all_tables.append(table_analysis)
                    table_id += 1

            except Exception as e:
                logger.error(f"Error analyzing region {region}: {str(e)}")
                all_tables.append({
                    "table_id": table_id,
                    "region": region,
                    "error": str(e),
                    "is_tabular": False
                })
                table_id += 1

        return {
            "sheet_name": sheet_name,
            "tables_found": len(all_tables),
            "tables": all_tables
        }

    def _analyze_single_table(self, df: pd.DataFrame, table_def: Dict[str, Any],
                             sheet: openpyxl.worksheet.worksheet.Worksheet,
                             table_id: int) -> Dict[str, Any]:
        """
        Analyze a single table within a region

        Args:
            df: DataFrame of the region
            table_def: Table definition from _detect_table_boundaries
            sheet: openpyxl worksheet
            table_id: Unique identifier for this table

        Returns:
            Dictionary with analysis results
        """
        analysis = {
            "table_id": table_id,
            "region": table_def['region'],
            "is_tabular": False,
            "confidence_score": 0.0,
            "indicators": {
                "has_header_row": False,
                "has_consistent_columns": False,
                "has_data_rows": False,
                "minimal_merged_cells": False,
                "rectangular_structure": False
            },
            "structure_info": {},
            "reasons": []
        }

        # Check 1: Has header row
        header_check = self._check_header_row(df)
        analysis["indicators"]["has_header_row"] = header_check["has_header"]
        if header_check["has_header"]:
            analysis["confidence_score"] += 25
            analysis["structure_info"]["detected_headers"] = header_check["headers"]
        else:
            analysis["reasons"].append("No clear header row detected")

        # Check 2: Consistent column structure
        column_consistency = self._check_column_consistency(df)
        analysis["indicators"]["has_consistent_columns"] = column_consistency["is_consistent"]
        if column_consistency["is_consistent"]:
            analysis["confidence_score"] += 25
        else:
            analysis["reasons"].append(column_consistency["reason"])

        # Check 3: Has data rows (not just headers)
        data_rows_check = self._check_data_rows(df)
        analysis["indicators"]["has_data_rows"] = data_rows_check["has_data"]
        if data_rows_check["has_data"]:
            analysis["confidence_score"] += 20
            analysis["structure_info"]["data_row_count"] = data_rows_check["row_count"]
        else:
            analysis["reasons"].append("No data rows found")

        # Check 4: Minimal merged cells in this specific region
        merged_cells_check = self._check_merged_cells_in_region(sheet, table_def['region'])
        analysis["indicators"]["minimal_merged_cells"] = merged_cells_check["is_minimal"]
        if merged_cells_check["is_minimal"]:
            analysis["confidence_score"] += 15
        else:
            analysis["reasons"].append(f"Too many merged cells ({merged_cells_check['count']})")

        analysis["structure_info"]["merged_cells_count"] = merged_cells_check["count"]

        # Check 5: Rectangular structure
        rectangular_check = self._check_rectangular_structure(df)
        analysis["indicators"]["rectangular_structure"] = rectangular_check["is_rectangular"]
        if rectangular_check["is_rectangular"]:
            analysis["confidence_score"] += 15
        else:
            analysis["reasons"].append(rectangular_check["reason"])

        # Determine if tabular based on confidence score
        if analysis["confidence_score"] >= 60:
            analysis["is_tabular"] = True
        else:
            analysis["is_tabular"] = False

        # Additional structure info
        analysis["structure_info"].update({
            "row_count": len(df),
            "column_count": len(df.columns),
            "fill_rate": round((df.notna().sum().sum() / (len(df) * len(df.columns))) * 100, 2) if len(df) > 0 and len(df.columns) > 0 else 0
        })

        return analysis

    def _check_merged_cells_in_region(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                                     region: Dict[str, int]) -> Dict[str, Any]:
        """
        Check merged cells within a specific region

        Args:
            sheet: openpyxl worksheet
            region: Region coordinates

        Returns:
            Dictionary with merged cell count and assessment
        """
        merged_count = 0

        for merged_range in sheet.merged_cells.ranges:
            # Check if merged range overlaps with region
            if (merged_range.min_row >= region['start_row'] and
                merged_range.max_row < region['end_row'] and
                merged_range.min_col >= region['start_col'] and
                merged_range.max_col < region['end_col']):
                merged_count += 1

        # Tables typically have few merged cells
        row_count = region['end_row'] - region['start_row']
        max_allowed = max(3, row_count * 0.1)

        return {
            "count": merged_count,
            "is_minimal": merged_count <= max_allowed
        }

    def _check_header_row(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Check if DataFrame has a clear header row"""
        if len(df.columns) == 0:
            return {"has_header": False, "headers": []}

        # Check if column names are meaningful (not Unnamed or generic)
        unnamed_count = sum(1 for col in df.columns if 'Unnamed' in str(col))

        # If more than 30% are unnamed, likely no proper header
        if unnamed_count / len(df.columns) > 0.3:
            return {"has_header": False, "headers": []}

        # Check if headers are not just numbers
        numeric_headers = sum(1 for col in df.columns if str(col).replace('.', '').replace('-', '').isdigit())

        if numeric_headers / len(df.columns) > 0.5:
            return {"has_header": False, "headers": []}

        return {
            "has_header": True,
            "headers": [str(col) for col in df.columns]
        }

    def _check_column_consistency(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Check if columns have consistent data types and structure"""
        if len(df) < 2:
            return {"is_consistent": False, "reason": "Insufficient rows for consistency check"}

        # Check if each column has a dominant data type
        consistent_columns = 0

        for col in df.columns:
            non_null = df[col].dropna()
            if len(non_null) == 0:
                continue

            # Get type distribution
            types = [type(v).__name__ for v in non_null]
            type_counts = pd.Series(types).value_counts()

            # If one type dominates (>70%), column is consistent
            if len(type_counts) > 0 and type_counts.iloc[0] / len(types) > 0.7:
                consistent_columns += 1

        consistency_rate = consistent_columns / len(df.columns) if len(df.columns) > 0 else 0

        if consistency_rate >= 0.7:
            return {"is_consistent": True, "consistency_rate": consistency_rate}
        else:
            return {
                "is_consistent": False,
                "reason": f"Low column consistency ({consistency_rate:.1%})",
                "consistency_rate": consistency_rate
            }

    def _check_data_rows(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Check if DataFrame has actual data rows"""
        if len(df) == 0:
            return {"has_data": False, "row_count": 0}

        # Count rows that have at least some non-null values
        non_empty_rows = df.notna().any(axis=1).sum()

        if non_empty_rows >= 1:  # At least one row with data
            return {"has_data": True, "row_count": non_empty_rows}
        else:
            return {"has_data": False, "row_count": 0}

    def _check_merged_cells(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
        """Check if sheet has minimal merged cells (indicator of tabular data)"""
        merged_count = len(sheet.merged_cells.ranges)

        # Tables typically have few merged cells
        # Allow up to 10% of rows to have merged cells
        max_allowed = max(5, sheet.max_row * 0.1)

        return {
            "count": merged_count,
            "is_minimal": merged_count <= max_allowed
        }

    def _check_rectangular_structure(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Check if data forms a rectangular structure"""
        if len(df) == 0 or len(df.columns) == 0:
            return {"is_rectangular": False, "reason": "Empty sheet"}

        # Calculate fill rate per row
        row_fill_rates = df.notna().sum(axis=1) / len(df.columns)

        # Check if most rows have similar fill rates (rectangular)
        if len(row_fill_rates) > 0:
            fill_rate_std = row_fill_rates.std()
            mean_fill_rate = row_fill_rates.mean()

            # If standard deviation is low and mean fill rate is reasonable, it's rectangular
            if fill_rate_std < 0.3 and mean_fill_rate > 0.3:
                return {"is_rectangular": True, "fill_rate_std": fill_rate_std}
            else:
                return {
                    "is_rectangular": False,
                    "reason": f"Irregular data distribution (std: {fill_rate_std:.2f})",
                    "fill_rate_std": fill_rate_std
                }
        else:
            return {"is_rectangular": False, "reason": "No data to analyze"}

    def _find_data_regions(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[Dict[str, int]]:
        """
        Find disconnected data regions in a sheet that could contain separate tables.
        Detects tables separated by blank rows AND/OR blank columns.

        Args:
            sheet: openpyxl worksheet

        Returns:
            List of regions as dictionaries with start_row, end_row, start_col, end_col
        """
        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row == 0 or max_col == 0:
            return []

        # Create a 2D grid to track cell occupancy
        grid = np.zeros((max_row, max_col), dtype=bool)

        # Mark cells with data
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=0):
            for col_idx, cell in enumerate(row, start=0):
                if cell.value is not None and str(cell.value).strip() != '':
                    grid[row_idx, col_idx] = True

        # Use flood fill algorithm to find connected components (regions)
        visited = np.zeros((max_row, max_col), dtype=bool)
        regions = []

        def flood_fill(start_row: int, start_col: int) -> Dict[str, int]:
            """Find connected region using flood fill with separate row/column gap thresholds"""
            stack = [(start_row, start_col)]
            min_row, max_row_local = start_row, start_row
            min_col, max_col_local = start_col, start_col

            while stack:
                r, c = stack.pop()

                if r < 0 or r >= max_row or c < 0 or c >= max_col:
                    continue
                if visited[r, c] or not grid[r, c]:
                    continue

                visited[r, c] = True
                min_row = min(min_row, r)
                max_row_local = max(max_row_local, r)
                min_col = min(min_col, c)
                max_col_local = max(max_col_local, c)

                # Check nearby cells with different thresholds for rows vs columns
                for dr in range(-self.blank_row_threshold, self.blank_row_threshold + 1):
                    for dc in range(-self.blank_col_threshold, self.blank_col_threshold + 1):
                        # Skip if both are zero (same cell) or if moving diagonally beyond threshold
                        if dr == 0 and dc == 0:
                            continue
                        # Allow gaps but prefer straight lines (not diagonal)
                        if abs(dr) + abs(dc) <= max(self.blank_row_threshold, self.blank_col_threshold) + 1:
                            nr, nc = r + dr, c + dc
                            if 0 <= nr < max_row and 0 <= nc < max_col:
                                if grid[nr, nc] and not visited[nr, nc]:
                                    stack.append((nr, nc))

            return {
                'start_row': min_row + 1,  # Convert to 1-indexed
                'end_row': max_row_local + 2,  # Inclusive, +1 for index, +1 for inclusive
                'start_col': min_col + 1,  # Convert to 1-indexed
                'end_col': max_col_local + 2  # Inclusive, +1 for index, +1 for inclusive
            }

        # Find all regions
        for row_idx in range(max_row):
            for col_idx in range(max_col):
                if grid[row_idx, col_idx] and not visited[row_idx, col_idx]:
                    region = flood_fill(row_idx, col_idx)
                    regions.append(region)

        # Post-process: Check for footer/summary rows after each region
        regions_with_footers = []
        for region in regions:
            extended_region = self._extend_region_with_footer(region, grid, max_row, max_col)
            regions_with_footers.append(extended_region)

        # Filter regions by minimum size
        valid_regions = []
        for region in regions_with_footers:
            row_count = region['end_row'] - region['start_row']
            col_count = region['end_col'] - region['start_col']

            if row_count >= self.min_table_rows and col_count >= self.min_table_cols:
                valid_regions.append(region)

        logger.info(f"Found {len(valid_regions)} potential table regions using flood fill")
        return valid_regions

    def _extend_region_with_footer(self, region: Dict[str, int], grid: np.ndarray,
                                   max_row: int, max_col: int) -> Dict[str, int]:
        """
        Check if there's a footer/summary row after a blank row gap and include it

        Args:
            region: Initial region boundaries
            grid: Grid of cell occupancy
            max_row: Maximum row in sheet
            max_col: Maximum column in sheet

        Returns:
            Extended region including footer if found
        """
        # Convert back to 0-indexed for grid operations
        region_end_row = region['end_row'] - 1  # Was already exclusive in 1-indexed
        region_start_col = region['start_col'] - 1
        region_end_col = region['end_col'] - 1

        # Look ahead for footer rows (skip up to blank_row_threshold + 1 rows)
        search_limit = min(region_end_row + self.blank_row_threshold + 3, max_row)

        footer_row_idx = None
        blank_rows_seen = 0

        for check_row_idx in range(region_end_row, search_limit):
            row_has_data = grid[check_row_idx, :].any()

            if not row_has_data:
                blank_rows_seen += 1
            else:
                # Found a row with data after blank row(s)
                if blank_rows_seen > 0 and blank_rows_seen <= self.blank_row_threshold + 1:
                    # Check if this row aligns with the table's column range
                    row_data_cols = np.where(grid[check_row_idx, :])[0]

                    if len(row_data_cols) > 0:
                        row_start_col = row_data_cols[0]
                        row_end_col = row_data_cols[-1]

                        # Check if footer row overlaps with table columns (allowing some flexibility)
                        table_col_range = set(range(region_start_col, region_end_col))
                        footer_col_range = set(range(row_start_col, row_end_col + 1))

                        overlap = len(table_col_range.intersection(footer_col_range))
                        overlap_ratio = overlap / len(table_col_range) if table_col_range else 0

                        # If footer overlaps significantly with table columns, include it
                        if overlap_ratio >= 0.5:
                            footer_row_idx = check_row_idx
                            logger.debug(f"Found footer row at {check_row_idx + 1} after {blank_rows_seen} blank rows")
                            # Continue checking for more footer rows
                            blank_rows_seen = 0
                        else:
                            # Not aligned, stop searching
                            break
                else:
                    # Too many blank rows or continuous data, stop
                    break

        # Extend region if footer found
        if footer_row_idx is not None:
            extended_region = region.copy()
            extended_region['end_row'] = footer_row_idx + 2  # +1 for 0->1 index, +1 for exclusive
            return extended_region

        return region

    def _detect_header_at_position(self, df: pd.DataFrame, row_idx: int) -> Dict[str, Any]:
        """
        Detect if a specific row could be a header row

        Args:
            df: DataFrame of the region
            row_idx: Row index to check

        Returns:
            Dictionary with detection results
        """
        if row_idx >= len(df):
            return {"is_header": False, "confidence": 0, "headers": []}

        row_data = df.iloc[row_idx]

        # Check if row has mostly text values
        text_count = sum(1 for val in row_data if isinstance(val, str))
        non_null_count = row_data.notna().sum()

        if non_null_count == 0:
            return {"is_header": False, "confidence": 0, "headers": []}

        text_ratio = text_count / non_null_count if non_null_count > 0 else 0

        # Check if values are unique (headers typically unique)
        unique_ratio = len(row_data.dropna().unique()) / non_null_count if non_null_count > 0 else 0

        # Calculate confidence
        confidence = (text_ratio * 0.6 + unique_ratio * 0.4) * 100

        is_header = confidence >= 50

        return {
            "is_header": is_header,
            "confidence": round(confidence, 2),
            "headers": [str(val) for val in row_data] if is_header else []
        }

    def _detect_table_boundaries(self, df: pd.DataFrame, region: Dict[str, int]) -> List[Dict[str, Any]]:
        """
        Detect individual table boundaries within a data region

        Args:
            df: DataFrame of the region
            region: Region coordinates

        Returns:
            List of table boundaries with metadata
        """
        if len(df) == 0:
            return []

        # For now, treat each region as one table, but check for sub-regions
        # Future enhancement: detect column alignment changes, sub-headers, etc.

        # Find potential header row (scan first 5 rows)
        header_row_idx = 0
        best_header_confidence = 0

        for idx in range(min(5, len(df))):
            header_check = self._detect_header_at_position(df, idx)
            if header_check['confidence'] > best_header_confidence:
                best_header_confidence = header_check['confidence']
                header_row_idx = idx

        # Create single table definition for this region
        table = {
            'region': region.copy(),
            'header_row': header_row_idx,
            'data_start_row': header_row_idx + 1,
            'estimated_header': best_header_confidence >= 50
        }

        return [table]

    def extract_tables(self, excel_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Extract all detected tables as DataFrames

        Args:
            excel_path: Path to Excel file
            sheet_name: Optional specific sheet

        Returns:
            Dictionary with extracted tables organized by sheet
        """
        analysis = self.analyse(excel_path, sheet_name)

        extracted_tables = {
            "file_path": excel_path,
            "sheets": []
        }

        for sheet_analysis in analysis['sheets_analyzed']:
            sheet_data = {
                "sheet_name": sheet_analysis['sheet_name'],
                "tables": []
            }

            if 'error' in sheet_analysis or 'message' in sheet_analysis:
                sheet_data['error'] = sheet_analysis.get('error') or sheet_analysis.get('message')
                extracted_tables['sheets'].append(sheet_data)
                continue

            for table in sheet_analysis.get('tables', []):
                if 'error' in table:
                    sheet_data['tables'].append({
                        'table_id': table['table_id'],
                        'error': table['error']
                    })
                    continue

                # Extract the DataFrame for this table
                region = table['region']
                try:
                    df = pd.read_excel(
                        excel_path,
                        sheet_name=sheet_analysis['sheet_name'],
                        header=None,
                        skiprows=region['start_row'] - 1,
                        nrows=region['end_row'] - region['start_row'],
                        usecols=range(region['start_col'] - 1, region['end_col'] - 1)
                    )

                    # Try to set proper headers if detected
                    if table['indicators'].get('has_header_row') and len(df) > 0:
                        # Use first row as header
                        df.columns = df.iloc[0]
                        df = df[1:].reset_index(drop=True)

                    sheet_data['tables'].append({
                        'table_id': table['table_id'],
                        'region': region,
                        'is_tabular': table['is_tabular'],
                        'confidence_score': table['confidence_score'],
                        'dataframe': df,
                        'shape': df.shape
                    })

                except Exception as e:
                    logger.error(f"Error extracting table {table['table_id']}: {str(e)}")
                    sheet_data['tables'].append({
                        'table_id': table['table_id'],
                        'region': region,
                        'error': str(e)
                    })

            extracted_tables['sheets'].append(sheet_data)

        return extracted_tables

    def extract_table_by_id(self, excel_path: str, sheet_name: str, table_id: int) -> Optional[pd.DataFrame]:
        """
        Extract a specific table by its ID

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name
            table_id: Table ID to extract

        Returns:
            DataFrame of the table, or None if not found
        """
        extracted = self.extract_tables(excel_path, sheet_name)

        for sheet_data in extracted['sheets']:
            if sheet_data['sheet_name'] == sheet_name:
                for table in sheet_data['tables']:
                    if table['table_id'] == table_id and 'dataframe' in table:
                        return table['dataframe']

        logger.warning(f"Table {table_id} not found in sheet '{sheet_name}'")
        return None

    def save_tables_to_files(self, excel_path: str, output_dir: str, sheet_name: Optional[str] = None,
                            file_format: str = 'csv') -> List[str]:
        """
        Save each detected table to separate files

        Args:
            excel_path: Path to Excel file
            output_dir: Directory to save extracted tables
            sheet_name: Optional specific sheet
            file_format: Output format ('csv', 'excel', 'json')

        Returns:
            List of saved file paths
        """
        import os
        from pathlib import Path

        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        extracted = self.extract_tables(excel_path, sheet_name)
        saved_files = []

        for sheet_data in extracted['sheets']:
            sheet_name_clean = sheet_data['sheet_name'].replace(' ', '_').replace('/', '_')

            for table in sheet_data['tables']:
                if 'dataframe' not in table:
                    continue

                table_id = table['table_id']
                df = table['dataframe']

                # Create filename
                base_name = f"{sheet_name_clean}_table_{table_id}"

                if file_format == 'csv':
                    file_path = output_path / f"{base_name}.csv"
                    df.to_csv(file_path, index=False)
                elif file_format == 'excel':
                    file_path = output_path / f"{base_name}.xlsx"
                    df.to_excel(file_path, index=False)
                elif file_format == 'json':
                    file_path = output_path / f"{base_name}.json"
                    df.to_json(file_path, orient='records', indent=2)
                else:
                    raise ValueError(f"Unsupported format: {file_format}")

                saved_files.append(str(file_path))
                logger.info(f"Saved table {table_id} to {file_path}")

        return saved_files

    def get_detection_summary(self, excel_path: str, sheet_name: Optional[str] = None) -> str:
        """
        Get a text summary of tabular structure detection (supports multiple tables per sheet)

        Args:
            excel_path: Path to Excel file
            sheet_name: Optional specific sheet

        Returns:
            Formatted text summary
        """
        analysis = self.analyse(excel_path, sheet_name)

        summary = []
        summary.append("=" * 70)
        summary.append("TABULAR STRUCTURE DETECTION - MULTI-TABLE SUPPORT")
        summary.append("=" * 70)
        summary.append(f"\nFile: {excel_path}")
        summary.append(f"\nSummary:")
        summary.append(f"  Total Sheets Analyzed: {analysis['summary']['total_sheets']}")
        summary.append(f"  Total Tables Found: {analysis['summary']['total_tables']}")
        summary.append(f"  Sheets with Tables: {analysis['summary']['sheets_with_tables']}")
        summary.append(f"  Sheets without Tables: {analysis['summary']['sheets_without_tables']}")

        for sheet_analysis in analysis['sheets_analyzed']:
            summary.append("\n" + "=" * 70)
            summary.append(f"SHEET: {sheet_analysis['sheet_name']}")
            summary.append("=" * 70)

            if 'error' in sheet_analysis:
                summary.append(f"  ERROR: {sheet_analysis['error']}")
                continue

            tables_found = sheet_analysis.get('tables_found', 0)
            summary.append(f"  Tables Found: {tables_found}")

            if 'message' in sheet_analysis:
                summary.append(f"  Message: {sheet_analysis['message']}")
                continue

            # Display each table
            for table in sheet_analysis.get('tables', []):
                summary.append("\n" + "-" * 70)
                summary.append(f"  TABLE #{table['table_id']}")
                summary.append("-" * 70)

                region = table['region']
                summary.append(f"    Region: Rows {region['start_row']}-{region['end_row']}, "
                             f"Cols {region['start_col']}-{region['end_col']}")

                if 'error' in table:
                    summary.append(f"    ERROR: {table['error']}")
                    continue

                summary.append(f"    Is Tabular: {'YES' if table['is_tabular'] else 'NO'}")
                summary.append(f"    Confidence Score: {table['confidence_score']}/100")

                summary.append(f"\n    Indicators:")
                for indicator, value in table['indicators'].items():
                    status = "✓" if value else "✗"
                    summary.append(f"      {status} {indicator.replace('_', ' ').title()}")

                if table.get('reasons'):
                    summary.append(f"\n    Reasons for Non-Tabular Classification:")
                    for reason in table['reasons']:
                        summary.append(f"      - {reason}")

                if table.get('structure_info'):
                    summary.append(f"\n    Structure Info:")
                    for key, value in table['structure_info'].items():
                        if key != 'detected_headers':  # Skip headers for brevity
                            summary.append(f"      - {key.replace('_', ' ').title()}: {value}")

                    # Show headers separately if present
                    if 'detected_headers' in table['structure_info']:
                        headers = table['structure_info']['detected_headers']
                        if len(headers) <= 10:
                            summary.append(f"      - Detected Headers: {', '.join(str(h) for h in headers)}")
                        else:
                            summary.append(f"      - Detected Headers: {', '.join(str(h) for h in headers[:10])}... ({len(headers)} total)")

        summary.append("\n" + "=" * 70)

        return "\n".join(summary)

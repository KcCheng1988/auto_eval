"""Field type classification and configuration recommendation"""

import pandas as pd
from typing import Dict, List, Any, Optional, Tuple
from collections import Counter
import inspect

from ..models.comparison_strategies.utils import NumericConverter, DateTimeConverter, is_null_like
from ..models.comparison_strategies.base import ComparisonStrategy
import sys
import importlib

class FieldType:
    """Enum-like class for field types"""
    NAME = "name"
    DATETIME = "datetime"
    DATE = "date"
    NUMERIC = "numeric"
    STRING = "string"
    ADDRESS = "address"
    UNKNOWN = "unknown"


class FieldClassifier:
    """
    Classify field types based on field names and values

    Analyzes a dataframe with field names and values to infer:
    - Field type (name, date, datetime, numeric, string, etc.)
    - Recommended comparison strategy
    - Suggested normalization settings
    """

    @staticmethod
    def _discover_strategies() -> Dict[str, List[str]]:
        """
        Automatically discover all available comparison strategies

        Returns:
            Dictionary mapping field types to list of strategy class names
        """
        # Import all strategy modules
        from ..models import comparison_strategies

        strategy_map = {
            FieldType.NAME: [],
            FieldType.DATETIME: [],
            FieldType.DATE: [],
            FieldType.NUMERIC: [],
            FieldType.STRING: [],
            FieldType.ADDRESS: [],
            FieldType.UNKNOWN: []
        }

        # Get all classes from comparison_strategies module
        for name, obj in inspect.getmembers(comparison_strategies, inspect.isclass):
            # Skip base classes
            if name in ['ComparisonStrategy', 'MatchResult']:
                continue

            # Check if it's a ComparisonStrategy subclass
            if issubclass(obj, ComparisonStrategy) and obj is not ComparisonStrategy:
                # Categorize by naming convention
                name_lower = name.lower()

                if 'name' in name_lower:
                    strategy_map[FieldType.NAME].append(name)
                elif 'datetime' in name_lower or 'timestamp' in name_lower:
                    strategy_map[FieldType.DATETIME].append(name)
                elif 'date' in name_lower and 'datetime' not in name_lower:
                    strategy_map[FieldType.DATE].append(name)
                elif 'numeric' in name_lower or 'number' in name_lower:
                    strategy_map[FieldType.NUMERIC].append(name)
                elif 'string' in name_lower or 'text' in name_lower:
                    strategy_map[FieldType.STRING].append(name)
                    # String strategies can also be used for addresses
                    strategy_map[FieldType.ADDRESS].append(name)
                else:
                    # Default to string if no specific category
                    strategy_map[FieldType.STRING].append(name)

        # Ensure each field type has at least one strategy
        for field_type in strategy_map:
            if not strategy_map[field_type]:
                # Fallback to string strategies
                strategy_map[field_type] = ['ExactStringMatch']

        return strategy_map

    def __init__(self):
        # Keywords for field name pattern matching
        self.name_keywords = [
            'name', 'customer', 'client', 'person', 'user', 'author',
            'contact', 'owner', 'manager', 'employee', 'staff'
        ]

        self.date_keywords = [
            'date', 'day', 'month', 'year', 'dob', 'birth',
            'created', 'updated', 'modified', 'expiry', 'expires'
        ]

        self.datetime_keywords = [
            'datetime', 'timestamp', 'time', 'created_at', 'updated_at',
            'modified_at', 'logged_at', 'recorded_at'
        ]

        self.numeric_keywords = [
            'amount', 'price', 'cost', 'fee', 'total', 'sum',
            'count', 'quantity', 'qty', 'number', 'size', 'age',
            'id', 'code', 'value', 'score', 'rate', 'percentage'
        ]

        self.address_keywords = [
            'address', 'street', 'city', 'state', 'country',
            'postal', 'zip', 'location', 'place'
        ]

    def classify_field(
        self,
        field_name: str,
        field_values: List[Any],
        sample_size: int = 100
    ) -> Dict[str, Any]:
        """
        Classify a single field based on name and values

        Args:
            field_name: Name of the field (will be cleaned)
            field_values: List of field values (will sample if too large)
            sample_size: Number of values to sample for analysis

        Returns:
            Dictionary with classification results
        """
        # Clean field name - remove leading/trailing whitespace and newlines
        if field_name is not None:
            field_name = str(field_name).strip().replace('\n', ' ').replace('\t', ' ').replace('\r', ' ')
            # Collapse multiple spaces into one
            import re
            field_name = re.sub(r'\s+', ' ', field_name)

        # Filter out null-like values first, before sampling
        non_null_values = [v for v in field_values if not is_null_like(v)]

        # Sample from non-null values if list is too large
        if len(non_null_values) > sample_size:
            import random
            sampled_values = random.sample(non_null_values, sample_size)
        else:
            sampled_values = non_null_values

        if not sampled_values:
            return self._create_classification_result(
                field_name=field_name,
                field_type=FieldType.UNKNOWN,
                confidence=0.0,
                reason="No non-empty values to analyze"
            )

        # Analyze field name
        field_name_lower = field_name.lower()
        name_hints = self._get_name_hints(field_name_lower)

        # Analyze field values
        value_type_counts = self._analyze_values(sampled_values)

        # Determine field type based on name hints and value analysis
        field_type, confidence, reason = self._determine_field_type(
            name_hints,
            value_type_counts,
            sampled_values
        )

        # Generate recommended strategy and settings
        strategy, settings = self._recommend_strategy(field_type, sampled_values)

        return self._create_classification_result(
            field_name=field_name,
            field_type=field_type,
            confidence=confidence,
            reason=reason,
            recommended_strategy=strategy,
            recommended_settings=settings,
            value_sample=sampled_values[:5]  # Include sample values
        )

    def _get_name_hints(self, field_name_lower: str) -> Dict[str, bool]:
        """Check field name for type hints"""
        hints = {
            'name': any(kw in field_name_lower for kw in self.name_keywords),
            'date': any(kw in field_name_lower for kw in self.date_keywords),
            'datetime': any(kw in field_name_lower for kw in self.datetime_keywords),
            'numeric': any(kw in field_name_lower for kw in self.numeric_keywords),
            'address': any(kw in field_name_lower for kw in self.address_keywords)
        }
        return hints

    def _analyze_values(self, values: List[Any]) -> Dict[str, int]:
        """Analyze value types in the sample"""
        type_counts = {
            'numeric': 0,
            'date': 0,
            'datetime': 0,
            'multi_word': 0,  # Likely names or addresses
            'single_word': 0
        }

        for value in values:
            # Try numeric conversion
            if NumericConverter.is_numeric(value):
                type_counts['numeric'] += 1
                continue

            # Try datetime conversion
            dt = DateTimeConverter.to_datetime(value)
            if dt is not None:
                # Check if it has time component
                if dt.hour != 0 or dt.minute != 0 or dt.second != 0:
                    type_counts['datetime'] += 1
                else:
                    type_counts['date'] += 1
                continue

            # Analyze as string
            value_str = str(value).strip()
            words = value_str.split()

            if len(words) >= 2:
                type_counts['multi_word'] += 1
            elif len(words) == 1:
                type_counts['single_word'] += 1
            # Note: len(words) == 0 shouldn't happen since we filter empty values
            # in classify_field(), but if it does, we just skip it

        return type_counts

    def _determine_field_type(
        self,
        name_hints: Dict[str, bool],
        value_type_counts: Dict[str, int],
        values: List[Any]
    ) -> Tuple[str, float, str]:
        """
        Determine field type based on name hints and value analysis

        Returns:
            Tuple of (field_type, confidence, reason)
        """
        total_values = sum(value_type_counts.values())

        if total_values == 0:
            return FieldType.UNKNOWN, 0.0, "No values to analyze"

        # Calculate percentages
        type_percentages = {
            k: v / total_values for k, v in value_type_counts.items()
        }

        # DateTime detection (check before date)
        if type_percentages.get('datetime', 0) > 0.8:
            confidence = type_percentages['datetime']
            if name_hints['datetime']:
                confidence = min(1.0, confidence + 0.1)
            return FieldType.DATETIME, confidence, f"{int(type_percentages['datetime']*100)}% datetime values"

        # Date detection
        if type_percentages.get('date', 0) > 0.8:
            confidence = type_percentages['date']
            if name_hints['date']:
                confidence = min(1.0, confidence + 0.1)
            return FieldType.DATE, confidence, f"{int(type_percentages['date']*100)}% date values"

        # Numeric detection
        if type_percentages.get('numeric', 0) > 0.8:
            confidence = type_percentages['numeric']
            if name_hints['numeric']:
                confidence = min(1.0, confidence + 0.1)
            return FieldType.NUMERIC, confidence, f"{int(type_percentages['numeric']*100)}% numeric values"

        # Name detection (multi-word strings with name hint)
        if name_hints['name'] and type_percentages.get('multi_word', 0) > 0.5:
            confidence = 0.7 + type_percentages['multi_word'] * 0.2
            return FieldType.NAME, confidence, "Name keyword + multi-word values"

        # Address detection
        if name_hints['address'] and type_percentages.get('multi_word', 0) > 0.6:
            confidence = 0.75
            return FieldType.ADDRESS, confidence, "Address keyword + multi-word values"

        # Default to string
        confidence = 0.6
        reason = "Default classification"

        if type_percentages.get('multi_word', 0) > 0.5:
            reason = f"{int(type_percentages['multi_word']*100)}% multi-word strings"
        elif type_percentages.get('single_word', 0) > 0.5:
            reason = f"{int(type_percentages['single_word']*100)}% single-word strings"

        return FieldType.STRING, confidence, reason

    def _recommend_strategy(
        self,
        field_type: str,
        values: List[Any]
    ) -> Tuple[str, Dict[str, Any]]:
        """
        Recommend comparison strategy and settings based on field type

        Returns:
            Tuple of (strategy_name, settings_dict)
        """
        if field_type == FieldType.NAME:
            return "ExactNameMatch", {
                "case_sensitive": False,
                "trim_whitespace": True,
                "normalize_unicode": True,
                "ignore_punctuation": False,
                "strip_line_breaks": True
            }

        elif field_type == FieldType.DATETIME:
            return "ExactDateTimeMatch", {}

        elif field_type == FieldType.DATE:
            return "DateOnlyMatch", {}

        elif field_type == FieldType.NUMERIC:
            # Check if values have decimals
            has_decimals = False
            for v in values[:20]:  # Sample check
                num = NumericConverter.to_float(v)
                if num is not None and num != int(num):
                    has_decimals = True
                    break

            return "ExactNumericMatch", {
                "decimal_precision": 2 if has_decimals else None
            }

        elif field_type == FieldType.STRING:
            return "ExactStringMatch", {
                "case_sensitive": False,
                "trim_whitespace": True,
                "normalize_unicode": True,
                "ignore_punctuation": False,
                "strip_line_breaks": True
            }

        elif field_type == FieldType.ADDRESS:
            return "ExactStringMatch", {
                "case_sensitive": False,
                "trim_whitespace": True,
                "normalize_unicode": True,
                "ignore_punctuation": False,
                "strip_line_breaks": True
            }

        else:  # UNKNOWN
            return "ExactStringMatch", {
                "case_sensitive": False,
                "trim_whitespace": True,
                "normalize_unicode": True,
                "ignore_punctuation": False,
                "strip_line_breaks": True
            }

    def _create_classification_result(
        self,
        field_name: str,
        field_type: str,
        confidence: float,
        reason: str,
        recommended_strategy: Optional[str] = None,
        recommended_settings: Optional[Dict[str, Any]] = None,
        value_sample: Optional[List[Any]] = None
    ) -> Dict[str, Any]:
        """Create a standardized classification result dictionary"""
        result = {
            'field_name': field_name,
            'field_type': field_type,
            'confidence': round(confidence, 2),
            'reason': reason,
            'recommended_strategy': recommended_strategy or 'ExactStringMatch',
            'recommended_settings': recommended_settings or {}
        }

        if value_sample:
            result['value_sample'] = value_sample

        return result

    def classify_dataframe(
        self,
        df: pd.DataFrame,
        field_name_col: str = 'field_name',
        field_value_col: str = 'field_value'
    ) -> pd.DataFrame:
        """
        Classify all fields in a dataframe

        Args:
            df: DataFrame with field names and values
            field_name_col: Column name containing field names
            field_value_col: Column name containing field values

        Returns:
            DataFrame with classification results
        """
        # Create a copy to avoid modifying original
        df_clean = df.copy()

        # Clean field names - remove leading/trailing whitespace and newlines
        import re
        df_clean[field_name_col] = df_clean[field_name_col].apply(
            lambda x: re.sub(r'\s+', ' ', str(x).strip().replace('\n', ' ').replace('\t', ' ').replace('\r', ' '))
            if x is not None else x
        )

        # Group by cleaned field name and aggregate values
        grouped = df_clean.groupby(field_name_col)[field_value_col].apply(list).reset_index()

        results = []
        for _, row in grouped.iterrows():
            field_name = row[field_name_col]
            field_values = row[field_value_col]

            classification = self.classify_field(field_name, field_values)

            # Start with base classification fields (excluding nested dicts and lists)
            result_row = {
                'field_name': classification['field_name'],
                'field_type': classification['field_type'],
                'confidence': classification['confidence'],
                'reason': classification['reason'],
                'recommended_strategy': classification['recommended_strategy']
            }

            # Flatten nested settings dict into separate columns with 'setting_' prefix
            settings = classification.get('recommended_settings', {})
            for key, value in settings.items():
                result_row[f'setting_{key}'] = value

            # Convert value_sample list to comma-separated string for Excel compatibility
            if 'value_sample' in classification:
                result_row['value_samples'] = ', '.join(str(v) for v in classification['value_sample'])

            results.append(result_row)

        return pd.DataFrame(results)

    def save_classification_to_excel(
        self,
        df: pd.DataFrame,
        output_path: str,
        field_name_col: str = 'field_name',
        field_value_col: str = 'field_value'
    ):
        """
        Classify fields and save results to Excel for manual review with dropdown menus

        Args:
            df: DataFrame with field names and values
            output_path: Path to save Excel file
            field_name_col: Column name containing field names
            field_value_col: Column name containing field values
        """
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.worksheet.table import Table, TableStyleInfo

        classification_df = self.classify_dataframe(df, field_name_col, field_value_col)

        # Define dropdown options for field types
        field_type_options = [
            FieldType.NAME,
            FieldType.DATETIME,
            FieldType.DATE,
            FieldType.NUMERIC,
            FieldType.STRING,
            FieldType.ADDRESS,
            FieldType.UNKNOWN
        ]

        # Automatically discover all available strategies
        strategy_map = self._discover_strategies()

        # Save to Excel with formatting and dropdowns
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            classification_df.to_excel(writer, sheet_name='Field Classification', index=False)

            # Get worksheet for formatting
            worksheet = writer.sheets['Field Classification']

            # Convert DataFrame to Excel Table (Ctrl+T equivalent)
            num_rows = len(classification_df)
            num_cols = len(classification_df.columns)

            # Define table range (A1 to last column and row)
            last_col_letter = chr(64 + num_cols)  # Works for A-Z
            if num_cols > 26:
                # Handle columns beyond Z (AA, AB, etc.)
                last_col_letter = chr(64 + (num_cols - 1) // 26) + chr(65 + (num_cols - 1) % 26)

            table_range = f"A1:{last_col_letter}{num_rows + 1}"

            # Create table with a nice style
            table = Table(displayName="FieldClassificationTable", ref=table_range)

            # Apply table style (like the default Ctrl+T style)
            style = TableStyleInfo(
                name="TableStyleMedium9",  # Blue table style
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style

            # Add table to worksheet
            worksheet.add_table(table)

            # Find column indices (Excel is 1-indexed)
            field_type_col_idx = list(classification_df.columns).index('field_type') + 1
            strategy_col_idx = list(classification_df.columns).index('recommended_strategy') + 1

            # Convert column index to Excel letter (A, B, C, ...)
            def col_index_to_letter(idx):
                """Convert 1-based column index to Excel letter"""
                letter = ''
                while idx > 0:
                    idx -= 1
                    letter = chr(65 + (idx % 26)) + letter
                    idx //= 26
                return letter

            field_type_col_letter = col_index_to_letter(field_type_col_idx)
            strategy_col_letter = col_index_to_letter(strategy_col_idx)

            # Add dropdown for field_type column
            field_type_dropdown = DataValidation(
                type="list",
                formula1=f'"{",".join(field_type_options)}"',
                allow_blank=False
            )
            field_type_dropdown.error = 'Invalid field type'
            field_type_dropdown.errorTitle = 'Invalid Entry'
            field_type_dropdown.prompt = 'Select field type'
            field_type_dropdown.promptTitle = 'Field Type'
            worksheet.add_data_validation(field_type_dropdown)

            # Apply field_type dropdown to all data rows
            num_rows = len(classification_df)
            field_type_range = f'{field_type_col_letter}2:{field_type_col_letter}{num_rows + 1}'
            field_type_dropdown.add(field_type_range)

            # Create a reference sheet with strategy options for each field type
            strategy_sheet = writer.book.create_sheet('Strategy_Options')

            # Write strategy options to reference sheet
            row_offset = 1
            for field_type, strategies in strategy_map.items():
                # Write field type as header
                strategy_sheet.cell(row=row_offset, column=1, value=field_type)

                # Write strategies below
                for i, strategy in enumerate(strategies, start=1):
                    strategy_sheet.cell(row=row_offset + i, column=1, value=strategy)

                # Define named range for this field type's strategies
                range_name = f"{field_type.upper().replace('-', '_')}_STRATEGIES"
                start_row = row_offset + 1
                end_row = row_offset + len(strategies)
                range_ref = f"Strategy_Options!$A${start_row}:$A${end_row}"

                # Add named range to workbook
                writer.book.create_named_range(range_name, strategy_sheet, range_ref)

                row_offset += len(strategies) + 2  # Leave gap between groups

            # Note: Dynamic dropdowns based on field_type require Excel formulas
            # For now, add a general dropdown with all strategies
            all_strategies = set()
            for strategies in strategy_map.values():
                all_strategies.update(strategies)

            strategy_dropdown = DataValidation(
                type="list",
                formula1=f'"{",".join(sorted(all_strategies))}"',
                allow_blank=False
            )
            strategy_dropdown.error = 'Invalid strategy'
            strategy_dropdown.errorTitle = 'Invalid Entry'
            strategy_dropdown.prompt = 'Select comparison strategy'
            strategy_dropdown.promptTitle = 'Comparison Strategy'
            worksheet.add_data_validation(strategy_dropdown)

            # Apply strategy dropdown to all data rows
            strategy_range = f'{strategy_col_letter}2:{strategy_col_letter}{num_rows + 1}'
            strategy_dropdown.add(strategy_range)

            # Add TRUE/FALSE dropdowns for all setting_* columns
            setting_columns = [col for col in classification_df.columns if col.startswith('setting_')]

            if setting_columns:
                # Create TRUE/FALSE dropdown
                bool_dropdown = DataValidation(
                    type="list",
                    formula1='"TRUE,FALSE"',
                    allow_blank=True  # Allow blank for optional settings
                )
                bool_dropdown.error = 'Please select TRUE or FALSE'
                bool_dropdown.errorTitle = 'Invalid Entry'
                bool_dropdown.prompt = 'Select TRUE or FALSE'
                bool_dropdown.promptTitle = 'Boolean Setting'
                worksheet.add_data_validation(bool_dropdown)

                # Apply to all setting columns
                for setting_col in setting_columns:
                    setting_col_idx = list(classification_df.columns).index(setting_col) + 1
                    setting_col_letter = col_index_to_letter(setting_col_idx)
                    setting_range = f'{setting_col_letter}2:{setting_col_letter}{num_rows + 1}'
                    bool_dropdown.add(setting_range)

            # Auto-adjust column widths
            for idx, col in enumerate(classification_df.columns, 1):
                max_length = max(
                    classification_df[col].astype(str).apply(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[col_index_to_letter(idx)].width = min(max_length + 2, 50)

            # Add instructions sheet
            instructions_sheet = writer.book.create_sheet('Instructions', 0)
            instructions = [
                ['Field Classification Instructions'],
                [''],
                ['How to use this file:'],
                ['1. Go to the "Field Classification" sheet'],
                ['2. Review the auto-detected field types and strategies'],
                ['3. Click on "field_type" cells to see dropdown menu with options:'],
                [f'   - {", ".join(field_type_options)}'],
                ['4. Click on "recommended_strategy" cells to see dropdown menu with available strategies'],
                ['5. Click on "setting_*" columns to see TRUE/FALSE dropdown menus'],
                ['6. Adjust numeric settings (e.g., decimal_precision, tolerance_seconds) manually'],
                ['7. Save the file when done'],
                ['7. Use FieldConfigLoader to load your edited configuration'],
                [''],
                ['Field Type -> Strategy Mapping:'],
            ]

            for field_type, strategies in strategy_map.items():
                instructions.append([f'{field_type}:', ', '.join(strategies)])

            for row_idx, row_data in enumerate(instructions, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    instructions_sheet.cell(row=row_idx, column=col_idx, value=cell_value)

            # Format instructions sheet
            instructions_sheet.column_dimensions['A'].width = 50
            instructions_sheet.column_dimensions['B'].width = 60

        print(f"Classification results saved to {output_path}")
        print(f"Total fields classified: {len(classification_df)}")
        print(f"\nField type distribution:")
        print(classification_df['field_type'].value_counts())
        print(f"\n✨ Excel enhancements:")
        print(f"  • Formatted as Excel Table with filters (Ctrl+T)")
        print(f"  • Dropdown menus for field_type and recommended_strategy")
        print(f"  • TRUE/FALSE dropdowns for all setting_* columns")
        print(f"  • Instructions sheet with usage guide")

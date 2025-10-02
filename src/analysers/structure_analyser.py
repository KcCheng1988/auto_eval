"""Sheet and structure analysis for Excel files"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Any
import logging

logger = logging.getLogger(__name__)


class StructureAnalyser:
    """Analyzes Excel file structure including sheets, images, and layout"""

    def __init__(self):
        self.workbook = None

    def analyse(self, excel_path: str) -> Dict[str, Any]:
        """
        Analyze Excel file structure

        Args:
            excel_path: Path to Excel file

        Returns:
            Dictionary containing structure analysis results
        """
        logger.info(f"Analyzing structure of: {excel_path}")

        self.workbook = openpyxl.load_workbook(excel_path, data_only=False)

        results = {
            "file_path": excel_path,
            "total_sheets": len(self.workbook.sheetnames),
            "active_sheets": [],
            "hidden_sheets": [],
            "sheet_details": [],
            "images_summary": {
                "total_images": 0,
                "sheets_with_images": []
            },
            "charts_summary": {
                "total_charts": 0,
                "sheets_with_charts": []
            }
        }

        # Analyze each sheet
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_info = self._analyse_sheet(sheet, sheet_name)
            results["sheet_details"].append(sheet_info)

            # Categorize sheets
            if sheet_info["is_hidden"]:
                results["hidden_sheets"].append(sheet_name)
            else:
                results["active_sheets"].append(sheet_name)

            # Aggregate images
            if sheet_info["images_count"] > 0:
                results["images_summary"]["total_images"] += sheet_info["images_count"]
                results["images_summary"]["sheets_with_images"].append(sheet_name)

            # Aggregate charts
            if sheet_info["charts_count"] > 0:
                results["charts_summary"]["total_charts"] += sheet_info["charts_count"]
                results["charts_summary"]["sheets_with_charts"].append(sheet_name)

        logger.info(f"Structure analysis complete: {results['total_sheets']} sheets analyzed")
        return results

    def _analyse_sheet(self, sheet: Worksheet, sheet_name: str) -> Dict[str, Any]:
        """
        Analyze individual sheet structure

        Args:
            sheet: Worksheet object
            sheet_name: Name of the sheet

        Returns:
            Dictionary with sheet analysis
        """
        info = {
            "name": sheet_name,
            "is_hidden": sheet.sheet_state == 'hidden',
            "dimensions": str(sheet.dimensions) if sheet.dimensions else "Empty",
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "images_count": 0,
            "charts_count": 0,
            "merged_cells_count": len(sheet.merged_cells.ranges),
            "has_merged_cells": len(sheet.merged_cells.ranges) > 0,
            "has_formulas": False,
            "has_data_validation": False,
            "has_conditional_formatting": False
        }

        # Count images
        if hasattr(sheet, '_images'):
            info["images_count"] = len(sheet._images)

        # Count charts
        if hasattr(sheet, '_charts'):
            info["charts_count"] = len(sheet._charts)

        # Check for formulas, data validation, conditional formatting
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                # Check for formulas
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    info["has_formulas"] = True

                # Check for data validation
                if cell.coordinate in sheet.data_validations.dataValidation:
                    info["has_data_validation"] = True

        # Check for conditional formatting
        if hasattr(sheet, 'conditional_formatting') and len(sheet.conditional_formatting) > 0:
            info["has_conditional_formatting"] = True

        return info

    def get_sheet_summary(self, excel_path: str) -> str:
        """
        Get a text summary of sheet structure

        Args:
            excel_path: Path to Excel file

        Returns:
            Formatted text summary
        """
        analysis = self.analyse(excel_path)

        summary = []
        summary.append("=" * 70)
        summary.append("EXCEL STRUCTURE ANALYSIS")
        summary.append("=" * 70)
        summary.append(f"\nFile: {excel_path}")
        summary.append(f"\nTotal Sheets: {analysis['total_sheets']}")
        summary.append(f"Active Sheets: {len(analysis['active_sheets'])}")
        summary.append(f"Hidden Sheets: {len(analysis['hidden_sheets'])}")

        if analysis['hidden_sheets']:
            summary.append(f"\nHidden Sheets List: {', '.join(analysis['hidden_sheets'])}")

        summary.append(f"\nTotal Images: {analysis['images_summary']['total_images']}")
        if analysis['images_summary']['sheets_with_images']:
            summary.append(f"Sheets with Images: {', '.join(analysis['images_summary']['sheets_with_images'])}")

        summary.append(f"\nTotal Charts: {analysis['charts_summary']['total_charts']}")
        if analysis['charts_summary']['sheets_with_charts']:
            summary.append(f"Sheets with Charts: {', '.join(analysis['charts_summary']['sheets_with_charts'])}")

        summary.append("\n" + "-" * 70)
        summary.append("SHEET DETAILS")
        summary.append("-" * 70)

        for sheet in analysis['sheet_details']:
            summary.append(f"\nSheet: {sheet['name']}")
            summary.append(f"  Status: {'Hidden' if sheet['is_hidden'] else 'Active'}")
            summary.append(f"  Dimensions: {sheet['dimensions']}")
            summary.append(f"  Size: {sheet['max_row']} rows x {sheet['max_column']} columns")
            summary.append(f"  Images: {sheet['images_count']}")
            summary.append(f"  Charts: {sheet['charts_count']}")
            summary.append(f"  Merged Cells: {sheet['merged_cells_count']}")
            summary.append(f"  Has Formulas: {sheet['has_formulas']}")
            summary.append(f"  Has Data Validation: {sheet['has_data_validation']}")
            summary.append(f"  Has Conditional Formatting: {sheet['has_conditional_formatting']}")

        summary.append("\n" + "=" * 70)

        return "\n".join(summary)

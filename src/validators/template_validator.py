"""Excel template validator for Task 1"""

import openpyxl
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional
import logging
import re

logger = logging.getLogger(__name__)


class ValidationResult:
    """Stores validation results and issues"""

    def __init__(self):
        self.is_valid = True
        self.errors = []
        self.warnings = []
        self.info = []

    def add_error(self, message: str):
        """Add an error (critical issue)"""
        self.errors.append(message)
        self.is_valid = False

    def add_warning(self, message: str):
        """Add a warning (non-critical issue)"""
        self.warnings.append(message)

    def add_info(self, message: str):
        """Add informational message"""
        self.info.append(message)

    def get_summary(self) -> Dict[str, Any]:
        """Get validation summary"""
        return {
            "is_valid": self.is_valid,
            "error_count": len(self.errors),
            "warning_count": len(self.warnings),
            "errors": self.errors,
            "warnings": self.warnings,
            "info": self.info
        }


class TemplateValidator:
    """Validates Excel templates against configuration rules"""

    def __init__(self, config: Dict[str, Any]):
        """
        Initialize validator with configuration

        Args:
            config: Template configuration from config_loader
        """
        self.config = config
        self.workbook = None
        self.result = ValidationResult()

    def validate_template(self, excel_path: str) -> ValidationResult:
        """
        Main validation method for Excel template

        Args:
            excel_path: Path to Excel file to validate

        Returns:
            ValidationResult object with validation results
        """
        logger.info(f"Starting validation for: {excel_path}")
        self.result = ValidationResult()

        try:
            # Load workbook
            self.workbook = openpyxl.load_workbook(excel_path, data_only=False)
            logger.info(f"Loaded workbook with {len(self.workbook.sheetnames)} sheets")

            # Validate sheet structure
            self._validate_sheets()

            # Extract and validate fields from each sheet
            self._validate_fields()

            # Validate data types and formats
            self._validate_data_quality()

            logger.info(f"Validation completed. Valid: {self.result.is_valid}")

        except Exception as e:
            logger.error(f"Error during validation: {str(e)}")
            self.result.add_error(f"Failed to validate template: {str(e)}")

        return self.result

    def _validate_sheets(self):
        """Validate that expected sheets exist"""
        expected_sheets = self.config.get("sheets", [])
        actual_sheets = self.workbook.sheetnames

        if not expected_sheets:
            self.result.add_info("No sheet validation rules configured")
            return

        for expected_sheet in expected_sheets:
            sheet_name = expected_sheet.get("name")
            is_required = expected_sheet.get("required", True)

            if sheet_name not in actual_sheets:
                if is_required:
                    self.result.add_error(f"Required sheet '{sheet_name}' not found")
                else:
                    self.result.add_warning(f"Optional sheet '{sheet_name}' not found")
            else:
                self.result.add_info(f"Sheet '{sheet_name}' found")

        # Check for unexpected sheets
        for actual_sheet in actual_sheets:
            expected_names = [s.get("name") for s in expected_sheets]
            if actual_sheet not in expected_names:
                self.result.add_warning(f"Unexpected sheet found: '{actual_sheet}'")

    def _validate_fields(self):
        """Extract and validate field-value pairs from sheets"""
        fields_config = self.config.get("fields", {})

        if not fields_config:
            self.result.add_info("No field validation rules configured")
            return

        for sheet_config in self.config.get("sheets", []):
            sheet_name = sheet_config.get("name")
            if sheet_name not in self.workbook.sheetnames:
                continue

            sheet = self.workbook[sheet_name]
            logger.info(f"Validating fields in sheet: {sheet_name}")

            # Extract field-value pairs using flexible detection
            extracted_fields = self._extract_fields_from_sheet(sheet, sheet_name)

            # Validate extracted fields against config
            sheet_fields = fields_config.get(sheet_name, {})
            for field_name, field_rules in sheet_fields.items():
                is_mandatory = field_rules.get("mandatory", False)

                if field_name not in extracted_fields:
                    if is_mandatory:
                        self.result.add_error(
                            f"Mandatory field '{field_name}' not found in sheet '{sheet_name}'"
                        )
                    else:
                        self.result.add_warning(
                            f"Optional field '{field_name}' not found in sheet '{sheet_name}'"
                        )
                else:
                    self.result.add_info(
                        f"Field '{field_name}' found in sheet '{sheet_name}'"
                    )

    def _extract_fields_from_sheet(
        self, sheet: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str
    ) -> Dict[str, Any]:
        """
        Extract field-value pairs from sheet using flexible pattern matching
        Handles form-like structures with fields at various locations

        Args:
            sheet: Openpyxl worksheet object
            sheet_name: Name of the sheet

        Returns:
            Dictionary of extracted field-value pairs
        """
        extracted = {}

        # Common patterns for field-value pairs:
        # 1. Field: Value (same row)
        # 2. Field in one cell, value in next cell
        # 3. Field in one row, value in row below

        for row in sheet.iter_rows():
            for idx, cell in enumerate(row):
                if cell.value is None:
                    continue

                cell_value = str(cell.value).strip()

                # Pattern 1: "Field: Value" or "Field = Value"
                if ':' in cell_value or '=' in cell_value:
                    parts = re.split(r'[:=]', cell_value, 1)
                    if len(parts) == 2:
                        field_name = parts[0].strip()
                        field_value = parts[1].strip()
                        extracted[field_name] = field_value
                        continue

                # Pattern 2: Field in current cell, value in next cell
                if idx + 1 < len(row) and row[idx + 1].value is not None:
                    # Check if current cell looks like a label (ends with :, ?, or is capitalized)
                    if (cell_value.endswith(':') or
                        cell_value.endswith('?') or
                        cell_value[0].isupper()):
                        field_name = cell_value.rstrip(':?').strip()
                        field_value = row[idx + 1].value
                        extracted[field_name] = field_value

        logger.debug(f"Extracted {len(extracted)} fields from sheet '{sheet_name}'")
        return extracted

    def _validate_data_quality(self):
        """Validate data types and formats of field values"""
        validation_rules = self.config.get("validation_rules", {})

        if not validation_rules:
            self.result.add_info("No data quality validation rules configured")
            return

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            extracted_fields = self._extract_fields_from_sheet(sheet, sheet_name)

            sheet_rules = validation_rules.get(sheet_name, {})
            for field_name, rules in sheet_rules.items():
                if field_name not in extracted_fields:
                    continue

                field_value = extracted_fields[field_name]
                self._validate_field_value(field_name, field_value, rules, sheet_name)

    def _validate_field_value(
        self, field_name: str, field_value: Any, rules: Dict[str, Any], sheet_name: str
    ):
        """
        Validate a single field value against rules

        Args:
            field_name: Name of the field
            field_value: Value to validate
            rules: Validation rules for this field
            sheet_name: Sheet containing the field
        """
        data_type = rules.get("type", "string")
        pattern = rules.get("pattern")
        allowed_values = rules.get("allowed_values")
        min_value = rules.get("min")
        max_value = rules.get("max")

        # Type validation
        if data_type == "integer":
            try:
                int(field_value)
            except (ValueError, TypeError):
                self.result.add_error(
                    f"Field '{field_name}' in sheet '{sheet_name}' should be integer, got: {field_value}"
                )
                return

        elif data_type == "float":
            try:
                float(field_value)
            except (ValueError, TypeError):
                self.result.add_error(
                    f"Field '{field_name}' in sheet '{sheet_name}' should be float, got: {field_value}"
                )
                return

        elif data_type == "date":
            # Openpyxl typically handles dates automatically
            if not isinstance(field_value, (pd.Timestamp, str)):
                self.result.add_error(
                    f"Field '{field_name}' in sheet '{sheet_name}' should be date format"
                )
                return

        # Pattern validation
        if pattern and isinstance(field_value, str):
            if not re.match(pattern, field_value):
                self.result.add_error(
                    f"Field '{field_name}' in sheet '{sheet_name}' does not match pattern: {pattern}"
                )

        # Allowed values validation
        if allowed_values and field_value not in allowed_values:
            self.result.add_error(
                f"Field '{field_name}' in sheet '{sheet_name}' has invalid value. "
                f"Allowed: {allowed_values}, Got: {field_value}"
            )

        # Range validation
        if min_value is not None:
            try:
                if float(field_value) < min_value:
                    self.result.add_error(
                        f"Field '{field_name}' in sheet '{sheet_name}' is below minimum: {min_value}"
                    )
            except (ValueError, TypeError):
                pass

        if max_value is not None:
            try:
                if float(field_value) > max_value:
                    self.result.add_error(
                        f"Field '{field_name}' in sheet '{sheet_name}' exceeds maximum: {max_value}"
                    )
            except (ValueError, TypeError):
                pass

    def get_extracted_data(self, excel_path: str) -> Dict[str, Dict[str, Any]]:
        """
        Extract all data from template for downstream processing

        Args:
            excel_path: Path to Excel file

        Returns:
            Dictionary with sheet names as keys and extracted data as values
        """
        workbook = openpyxl.load_workbook(excel_path, data_only=False)
        all_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            all_data[sheet_name] = self._extract_fields_from_sheet(sheet, sheet_name)

        return all_data
